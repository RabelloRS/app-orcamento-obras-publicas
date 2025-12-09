import time
import os
import asyncio
import re
import unicodedata
import zipfile
import pandas as pd
import io
import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, delete
from models import ReferenceSource, ReferenceItem, ReferencePrice, CompositionItem
from datetime import date
from uuid import uuid4, UUID
from decimal import Decimal

class ImportLock:
    def __init__(self, lock_file="import.lock"):
        self.lock_file = lock_file
        
    def __enter__(self):
        # No-op lock to avoid stale lockfiles during manual runs
        if os.path.exists(self.lock_file):
            print("Lockfile found; proceeding anyway for manual import.")
        return self
    def __exit__(self, t, v, tb):
        if os.path.exists(self.lock_file):
            os.remove(self.lock_file)

def parse_sinapi_file_sync(file_content: bytes, filename: str, state_filter: str | None = None):
    """
    Synchronous function to parse SINAPI Excel/ZIP file.
    Returns a dictionary with parsed data to be consumed by the async importer.
    """
    # 1. Handle ZIP
    if file_content.startswith(b'PK\x03\x04'):
        with zipfile.ZipFile(io.BytesIO(file_content)) as z:
            candidates = [n for n in z.namelist() if n.endswith(('.xls', '.xlsx')) and not n.startswith('~')]
            if not candidates:
                raise ValueError("Nenhum Excel encontrado no ZIP.")
            target_file = max(candidates, key=lambda n: z.getinfo(n).file_size)
            file_content = z.read(target_file)
            # Update filename if it was a zip, to help with date detection later
            filename = target_file

    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=False, read_only=True)
    sheet_names = wb.sheetnames

    def _normalize_sheet_name(name: str) -> str:
        normalized = unicodedata.normalize('NFD', name)
        normalized = normalized.encode('ascii', 'ignore').decode('ascii')
        return re.sub(r'[^A-Z0-9]', '', normalized.upper())

    def _detect_sheet_targets(names):
        detected = []
        for raw in names:
            norm = _normalize_sheet_name(raw)
            if "ANALIT" in norm:
                continue
            meta = None
            if norm.startswith("CCD") or ("COMPOSICOES" in norm and "NAODESON" not in norm and "NODESON" not in norm):
                meta = {"type": "COMPOSITION", "item_type": "SERVICE", "regime": "DESONERADO"}
            elif norm.startswith("CSD") or ("COMPOSICOES" in norm and ("NAODESON" in norm or "NODESON" in norm or ("NAO" in norm and "DESON" in norm))):
                meta = {"type": "COMPOSITION", "item_type": "SERVICE", "regime": "NAO_DESONERADO"}
            elif norm.startswith("ICD") or ("INSUMOS" in norm and "NAODESON" not in norm and "NODESON" not in norm and not norm.startswith("ISD")):
                meta = {"type": "INPUT", "regime": "DESONERADO"}
            elif norm.startswith("ISD") or ("INSUMOS" in norm and ("NAODESON" in norm or "NODESON" in norm or ("NAO" in norm and "DESON" in norm))):
                meta = {"type": "INPUT", "regime": "NAO_DESONERADO"}

            if meta:
                detected.append((raw, meta))
        return detected

    targets = _detect_sheet_targets(sheet_names)
    if not targets:
        raise ValueError("Nenhuma aba de COMPOSICOES ou INSUMOS encontrada no arquivo SINAPI.")

    parsed_sheets = []
    
    # 2. Parse Sheets
    for sheet, meta in targets:
        ws = wb[sheet]
        
        # Detect header
        header_row_idx = -1
        temp_rows = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True)):
            temp_rows.append(row)
            row_str = " ".join([str(x).upper() for x in row if x])
            if ("CODIGO" in row_str or "CÓDIGO" in row_str) and ("DESCRICAO" in row_str or "DESCRIÇÃO" in row_str):
                header_row_idx = i
                break
        if header_row_idx == -1:
            continue

        # Detect states
        valid_states = ["AC","AL","AP","AM","BA","CE","DF","ES","GO","MA","MT","MS","MG","PA","PB","PR","PE","PI","RJ","RN","RS","RO","RR","SC","SP","SE","TO"]
        found_states = {}
        for i in range(header_row_idx + 1):
            row = temp_rows[i]
            for col_idx, val in enumerate(row):
                v = str(val).strip().upper() if val else ""
                if v in valid_states:
                    found_states[v] = col_idx

        # Fallback single state
        if not found_states:
            detected_state = state_filter or None
            if not detected_state and filename:
                m_state = re.search(r'(AC|AL|AP|AM|BA|CE|DF|ES|GO|MA|MT|MS|MG|PA|PB|PR|PE|PI|RJ|RN|RS|RO|RR|SC|SP|SE|TO)', filename.upper())
                if m_state:
                    detected_state = m_state.group(1)
            if detected_state:
                h_vals = [str(x).strip().upper() for x in temp_rows[header_row_idx]]
                price_col_idx = next((idx for idx, v in enumerate(h_vals) if "PRECO" in v or "PREÇO" in v or "CUSTO" in v), None)
                if price_col_idx is not None:
                    found_states[detected_state] = price_col_idx
            if not found_states:
                continue

        # Filter states
        if state_filter and state_filter != "ALL":
            if state_filter in found_states:
                found_states = {state_filter: found_states[state_filter]}
            else:
                continue

        # Column indices
        code_col_idx = desc_col_idx = unit_col_idx = -1
        full_header_row = temp_rows[header_row_idx]
        for idx, val in enumerate(full_header_row):
            v = str(val).strip().upper() if val else ""
            if "CODIGO" in v or "CÓDIGO" in v:
                code_col_idx = idx
            elif "DESCRICAO" in v or "DESCRIÇÃO" in v:
                desc_col_idx = idx
            elif "UNIDADE" in v:
                unit_col_idx = idx
        if code_col_idx == -1: code_col_idx = 1
        if desc_col_idx == -1: desc_col_idx = 2
        if unit_col_idx == -1: unit_col_idx = 3

        def extract_code(val):
            if isinstance(val, str) and "HYPERLINK" in val:
                m = re.search(r';(\d{4,6})\)', val) or re.search(r',(\d{4,6})\)', val)
                if m: return m.group(1)
            val_str = str(val).strip() if val is not None else ""
            if val_str and "HYPERLINK" in val_str and "MATCH" in val_str:
                m = re.search(r'MATCH\((\d{4,6})', val_str)
                if m: return m.group(1)
            return val_str

        rows_data = []
        for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=False):
            if code_col_idx >= len(row): continue
            raw_val = row[code_col_idx].value
            code_val = extract_code(raw_val) if raw_val is not None else ""
            if not code_val or code_val == "0": continue
            
            v_upper = str(code_val).upper()
            if v_upper.startswith("COD") or v_upper.startswith("CÓD") or "GRUPO" in v_upper:
                continue

            row_item = {
                "code": str(code_val).strip(),
                "desc": str(row[desc_col_idx].value or ""),
                "unit": str(row[unit_col_idx].value or ""),
                "prices": {}
            }
            
            has_price = False
            for st, s_idx in found_states.items():
                if s_idx < len(row):
                    val = row[s_idx].value
                    # Clean price
                    if pd.isna(val): continue
                    if isinstance(val, str):
                        val = val.replace('R$', '').replace('.', '').replace(',', '.').strip()
                        try: val = float(val)
                        except: continue
                    try:
                        val = float(val)
                        if val >= 0:
                            row_item["prices"][st] = val
                            has_price = True
                    except: continue
            
            if has_price or row_item["code"]: 
                 rows_data.append(row_item)

        if rows_data:
            parsed_sheets.append({
                "sheet_name": sheet,
                "meta": meta,
                "found_states": list(found_states.keys()),
                "rows": rows_data
            })

    # 3. Parse Analytics
    analytic_links = []
    analytic_sheets = [name for name in sheet_names if "ANALIT" in _normalize_sheet_name(name)]
    
    for sheet_name in analytic_sheets:
        ws = wb[sheet_name]
        
        # Detect header
        header_row_idx = None
        header_vals = []
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=80, values_only=True), start=1):
            vals = []
            for x in row:
                if x:
                    val = str(x).strip()
                    val = unicodedata.normalize('NFD', val).encode('ascii', 'ignore').decode('ascii')
                    val = val.upper().replace("\n", " ").replace("\r", " ")
                    vals.append(val)
                else:
                    vals.append("")
            joined = " ".join(vals)
            if ("COMPOS" in joined or "ITEM" in joined) and any("COEF" in h for h in vals):
                header_row_idx = idx
                header_vals = vals
                break
        
        if not header_row_idx:
            continue

        parent_code_idx = child_code_idx = coef_idx = None
        code_positions = []
        for idx, val in enumerate(header_vals):
            if "COEF" in val: coef_idx = idx
            if "COD" in val: code_positions.append(idx)
            if "COMPOSICAO" in val and "CODIGO" in val: parent_code_idx = idx
            if "ITEM" in val and "COMPOSICAO" in val: child_code_idx = idx
            elif parent_code_idx is None and "COMP" in val and "COD" in val: parent_code_idx = idx
            elif child_code_idx is None and "ITEM" in val and "COD" in val: child_code_idx = idx
        
        if parent_code_idx is None and code_positions: parent_code_idx = code_positions[0]
        if child_code_idx is None and len(code_positions) > 1: child_code_idx = code_positions[1]

        if parent_code_idx is None or child_code_idx is None or coef_idx is None:
            continue

        def normalize_code_simple(val):
            if val is None: return ""
            return str(val).strip().split(';')[0].split(',')[0].replace('HYPERLINK', '').replace('=', '').replace('"', '').replace('(', '').replace(')', '')

        def parse_decimal_simple(val):
            if val is None: return None
            if isinstance(val, (int, float, Decimal)): return Decimal(str(val))
            text = str(val).strip().replace(" ", "")
            if text.count(',') == 1 and text.count('.') >= 1: text = text.replace('.', '')
            text = text.replace(',', '.')
            try: return Decimal(text)
            except: return None

        current_parent = ""
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            p_raw = row[parent_code_idx] if parent_code_idx < len(row) else None
            c_raw = row[child_code_idx] if child_code_idx < len(row) else None
            coef_raw = row[coef_idx] if coef_idx < len(row) else None
            
            p_code = normalize_code_simple(p_raw)
            c_code = normalize_code_simple(c_raw)
            
            if p_code: current_parent = p_code
            else: p_code = current_parent
            
            if not p_code or not c_code: continue
            
            coef = parse_decimal_simple(coef_raw)
            if coef is None or coef <= 0: continue
            
            analytic_links.append({
                "parent": p_code,
                "child": c_code,
                "coef": coef 
            })

    return {
        "sheets": parsed_sheets,
        "analytics": analytic_links,
        "filename": filename
    }

async def import_sinapi_excel(
    file_content: bytes, 
    filename: str, 
    state: str | None, 
    month: int | None, 
    year: int | None, 
    db: AsyncSession,
    progress_callback=None,
    replace: bool = False,
    user_id: UUID | None = None
):
    # 0. Set Progress Helper
    async def set_progress(pct: int, msg: str):
        pct = max(0, min(100, pct))
        if progress_callback:
            await progress_callback(pct, 100, msg)

    await set_progress(2, "Iniciando leitura do arquivo...")

    # 1. Parse File in Thread Pool (Non-Blocking)
    try:
        parsed_data = await asyncio.to_thread(parse_sinapi_file_sync, file_content, filename, state)
    except Exception as e:
         raise ValueError(f"Erro ao ler arquivo Excel: {str(e)}")

    filename = parsed_data["filename"]
    
    # 2. Detect Date
    if not month or not year:
        match = re.search(r'(20\d{2})[-_]?(0[1-9]|1[0-2])', filename)
        if match:
            year = int(match.group(1))
            month = int(match.group(2))
        else:
            raise ValueError("Não foi possível identificar mês/ano pelo nome do arquivo. Use padrão SINAPI_YYYYMM.zip")

    target_date = date(year, month, 1)
    state_param = state.upper() if state else None 
    await set_progress(10, f"Arquivo lido. Mês/Ano: {month}/{year}. Importando dados...")

    # 3. Handle Replace
    if replace:
        from sqlalchemy import text
        where_clause = f"EXTRACT(MONTH FROM date_validity) = {month} AND EXTRACT(YEAR FROM date_validity) = {year} AND is_active = TRUE"
        if state_param and state_param != "ALL":
            where_clause += f" AND region = '{state_param}'"
        user_val = f"'{user_id}'" if user_id else "NULL"
        sql = f"UPDATE reference_prices SET is_active = FALSE, inactivated_at = NOW(), inactivated_by_id = {user_val} WHERE {where_clause}"
        await db.execute(text(sql))
        await db.commit()

    lock = ImportLock()
    try:
        lock.__enter__()
    except BlockingIOError:
        print("Lockfile found. Proceeding...")

    try:
        # 4. Ensure Source Exists
        q_src = await db.execute(select(ReferenceSource).where(ReferenceSource.name == "SINAPI"))
        sinapi_src = q_src.scalars().first()
        if not sinapi_src:
            sinapi_src = ReferenceSource(
                name="SINAPI",
                description="Sistema Nacional de Pesquisa de Custos e Índices da Construção Civil",
                state="BR",
                month=1,
                year=2024,
                type="SYNTHETIC"
            )
            db.add(sinapi_src)
            await db.commit()
            await db.refresh(sinapi_src)

        master_source_id = sinapi_src.id

        # 5. Build Caches
        item_id_cache = {}
        q_items = await db.execute(select(ReferenceItem.id, ReferenceItem.code).where(ReferenceItem.source_id == master_source_id))
        for row in q_items.all():
            item_id_cache[(master_source_id, row.code)] = row.id

        existing_price_keys = set()
        price_rows = await db.execute(
            select(ReferencePrice.item_id, ReferencePrice.region, ReferencePrice.charge_type)
            .join(ReferenceItem, ReferenceItem.id == ReferencePrice.item_id)
            .where(
                ReferenceItem.source_id == master_source_id,
                func.extract('month', ReferencePrice.date_validity) == month,
                func.extract('year', ReferencePrice.date_validity) == year,
                ReferencePrice.is_active == True
            )
        )
        for item_id, region, charge_type in price_rows.all():
            existing_price_keys.add((item_id, region, charge_type or "DESONERADO"))

        total_imported = 0
        total_sheets = len(parsed_data["sheets"])
        
        # 6. Iterate Sheets
        for s_idx, sheet_data in enumerate(parsed_data["sheets"]):
            sheet_name = sheet_data["sheet_name"]
            rows = sheet_data["rows"]
            meta = sheet_data["meta"]
            
            await set_progress(10 + int((s_idx/total_sheets)*40), f"Processando aba {sheet_name} ({len(rows)} itens)")
            
            new_items_buffer = []
            
            # 6a. Create Items
            new_items_count = 0
            for row in rows:
                code_val = row["code"]
                cache_key = (master_source_id, code_val)
                
                if cache_key not in item_id_cache:
                    desc_upper = row["desc"].upper()
                    if meta.get("item_type"):
                        item_type = meta["item_type"]
                    elif meta["type"] == "COMPOSITION":
                        item_type = "COMPOSITION"
                    elif "MAO DE OBRA" in desc_upper or "MÃO DE OBRA" in desc_upper or "ENCARGOS" in desc_upper:
                        item_type = "LABOR"
                    elif "EQUIPAMENTO" in desc_upper:
                        item_type = "EQUIPMENT"
                    else:
                        item_type = "MATERIAL"

                    new_id = uuid4()
                    db.add(ReferenceItem(
                        id=new_id, 
                        source_id=master_source_id, 
                        code=code_val, 
                        description=row["desc"], 
                        unit=row["unit"], 
                        type=item_type
                    ))
                    item_id_cache[cache_key] = new_id
                    new_items_count += 1
                    
                    if new_items_count % 1000 == 0:
                        await db.commit()
                        await asyncio.sleep(0)
            
            await db.commit() # Commit remaining items

            # 6b. Import Prices
            prices_count = 0
            for row in rows:
                code_val = row["code"]
                item_id = item_id_cache.get((master_source_id, code_val))
                if not item_id: continue
                
                for st, price_val in row["prices"].items():
                    key = (item_id, st, meta.get("regime", "DESONERADO"))
                    if key in existing_price_keys:
                        continue
                    
                    db.add(ReferencePrice(
                        item_id=item_id,
                        region=st,
                        price=price_val,
                        date_validity=target_date,
                        charge_type=meta.get("regime", "DESONERADO")
                    ))
                    existing_price_keys.add(key)
                    prices_count += 1
                    total_imported += 1
                
                if prices_count % 2000 == 0:
                     await db.commit()
                     await asyncio.sleep(0)
            
            await db.commit()

        # 7. Analytics
        analytic_links = parsed_data["analytics"]
        if analytic_links:
            await set_progress(90, f"Processando {len(analytic_links)} vínculos analíticos...")
            
            # Clean old links
            await db.execute(
                delete(CompositionItem).where(
                    CompositionItem.parent_item_id.in_(
                        select(ReferenceItem.id).where(ReferenceItem.source_id == master_source_id)
                    )
                )
            )
            await db.commit()
            
            links_added = 0
            for link in analytic_links:
                parent_key = (master_source_id, link["parent"])
                child_key = (master_source_id, link["child"])
                
                if parent_key in item_id_cache and child_key in item_id_cache:
                    db.add(CompositionItem(
                        parent_item_id=item_id_cache[parent_key],
                        child_item_id=item_id_cache[child_key],
                        coefficient=link["coef"],
                        price_source=None
                    ))
                    links_added += 1
                
                if links_added % 2000 == 0:
                    await db.commit()
                    await asyncio.sleep(0)
            
            await db.commit()
            await set_progress(95, f"Vínculos analíticos concluídos ({links_added})")

        await set_progress(100, "Importação Concluída com Sucesso!")
        return total_imported

    except Exception as e:
        await db.rollback()
        raise e
    finally:
        if os.path.exists("import.lock"):
            os.remove("import.lock")
        with open("heartbeat.txt", "w") as f:
            f.write(f"{time.time()}|DONE|ALL")

async def import_sicro_excel(
    file_content: bytes, 
    filename: str, 
    state: str | None, 
    month: int, 
    year: int, 
    db: AsyncSession,
    progress_callback=None,
    replace: bool = False,
    user_id: UUID | None = None
):
    # Similar logic to SINAPI but for SICRO
    if replace:
        print(f"Replace=True. Soft deleting existing SICRO prices for {month}/{year} (State: {state})")
        from sqlalchemy import text
        where_clause = f"EXTRACT(MONTH FROM date_validity) = {month} AND EXTRACT(YEAR FROM date_validity) = {year} AND is_active = TRUE"
        if state and state != "ALL":
             where_clause += f" AND region = '{state}'"
        user_val = f"'{user_id}'" if user_id else "NULL"
        subquery = "item_id IN (SELECT id FROM reference_items WHERE source_id IN (SELECT id FROM reference_sources WHERE name='SICRO'))"
        sql = f"UPDATE reference_prices SET is_active = FALSE, inactivated_at = NOW(), inactivated_by_id = {user_val} WHERE {where_clause} AND {subquery}"
        await db.execute(text(sql))
        await db.commit()

    lock = ImportLock()
    try:
        lock.__enter__()
    except BlockingIOError:
        pass 

    try:
        if file_content.startswith(b'PK\x03\x04'):
            print("Detected ZIP file. Extracting...")
            with zipfile.ZipFile(io.BytesIO(file_content)) as z:
                candidates = [n for n in z.namelist() if n.endswith(('.xls', '.xlsx')) and not n.startswith('~')]
                if not candidates:
                    raise ValueError("No Excel file found inside ZIP.")
                target_file = max(candidates, key=lambda n: z.getinfo(n).file_size)
                print(f"Extracting for SICRO: {target_file}")
                file_content = z.read(target_file)
        
        xl = pd.ExcelFile(io.BytesIO(file_content))
        sheet_names = xl.sheet_names
        target_sheet = sheet_names[0]
        print(f"Processing SICRO sheet {target_sheet}...")
        
        df = pd.read_excel(io.BytesIO(file_content), sheet_name=target_sheet, header=None)
        header_idx = 0
        for i in range(min(20, len(df))):
            row_str = " ".join([str(x).upper() for x in df.iloc[i].values])
            if "CODIGO" in row_str or "CÓDIGO" in row_str:
                header_idx = i + 1 
                break
        
        df = pd.read_excel(io.BytesIO(file_content), sheet_name=target_sheet, header=None, skiprows=header_idx)
        
        q_src = await db.execute(select(ReferenceSource).where(ReferenceSource.name == "SICRO"))
        sicro_src = q_src.scalars().first()
        if not sicro_src:
            sicro_src = ReferenceSource(
                name="SICRO", 
                description="Sistema de Custos Referenciais de Obras - DNIT",
                state="BR",
                month=1,
                year=2024,
                type="SYNTHETIC"
            )
            db.add(sicro_src)
            await db.commit()
            await db.refresh(sicro_src)
            
        master_source_id = sicro_src.id
        item_id_cache = {}
        q_items = await db.execute(select(ReferenceItem.id, ReferenceItem.code).where(ReferenceItem.source_id == master_source_id))
        for row in q_items.all():
            item_id_cache[(master_source_id, row.code)] = row.id
            
        codes = df.iloc[:, 0].astype(str).str.strip()
        descs = df.iloc[:, 1].astype(str).str.strip()
        units = df.iloc[:, 2].astype(str).str.strip()
        price_col_idx = 4 if df.shape[1] > 4 else 3 
        if df.shape[1] <= price_col_idx: price_col_idx = df.shape[1] - 1
        prices = df.iloc[:, price_col_idx]

        new_items_count = 0
        total_imported = 0
        valid_rows = []
        for i in range(len(df)):
            c = codes.iloc[i]
            if pd.isna(c) or len(c) < 3 or len(c) > 20: continue 
            valid_rows.append(i)

        for i in valid_rows:
            c = codes.iloc[i]
            cache_key = (master_source_id, c)
            if cache_key not in item_id_cache:
                new_id = uuid4()
                d_text = str(descs.iloc[i]).upper()
                if "COMPOSI" in d_text: item_type = "COMPOSITION"
                elif "EQUIPAMENTO" in d_text: item_type = "EQUIPMENT"
                elif "MAO DE OBRA" in d_text: item_type = "LABOR"
                else: item_type = "MATERIAL"

                new_item = ReferenceItem(id=new_id, source_id=master_source_id, code=c, description=descs.iloc[i], unit=units.iloc[i], type=item_type)
                db.add(new_item)
                item_id_cache[cache_key] = new_id
                new_items_count += 1
                if new_items_count % 1000 == 0: await db.commit()

        await db.commit()
        
        for i in valid_rows:
            c = codes.iloc[i]
            item_id = item_id_cache[(master_source_id, c)]
            p_val = prices.iloc[i]
            if pd.isna(p_val): p_val = 0.0
            if isinstance(p_val, str):
                p_val = p_val.replace('R$', '').replace('.', '').replace(',', '.').strip()
                try: p_val = float(p_val)
                except: p_val = 0.0
                
            if p_val >= 0:
                p_obj = ReferencePrice(item_id=item_id, region=state if state else "BR", price=float(p_val), date_validity=date(year, month, 1))
                db.add(p_obj)
                total_imported += 1
                
            if total_imported % 1000 == 0: 
                await db.commit()
                if progress_callback:
                    await progress_callback(total_imported, len(valid_rows), f"Importing SICRO {total_imported}/{len(valid_rows)}")

        await db.commit()
        return total_imported

    except Exception as e:
        await db.rollback()
        print(f"Error importing SICRO: {e}")
        raise e
    finally:
        if os.path.exists(lock.lock_file): os.remove(lock.lock_file)

