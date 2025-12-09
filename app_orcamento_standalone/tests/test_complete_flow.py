#!/usr/bin/env python3
"""
Test the complete flow:
1. Import SINAPI with compositions
2. Check compositions are returned by API endpoint
"""
import asyncio
import requests
import time
from io import BytesIO
import zipfile
from openpyxl import Workbook
from services.importer import import_sinapi_excel
from database import AsyncSessionLocal
from sqlalchemy import text

async def create_test_file():
    """Create test SINAPI file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "CCD-RS-12-2025"
    
    # Headers
    ws.cell(row=1, column=1).value = "CÓDIGO"
    ws.cell(row=1, column=2).value = "DESCRIÇÃO"
    ws.cell(row=1, column=3).value = "UNIDADE"
    ws.cell(row=1, column=4).value = "RS"
    
    # Items
    items = [
        ("50000", "Serviço A", "UN", 500.00),
        ("50001", "Serviço B", "M2", 300.00),
        ("60000", "Insumo 1", "KG", 25.00),
        ("60001", "Insumo 2", "L", 15.00),
    ]
    
    for row_idx, (code, desc, unit, price) in enumerate(items, 2):
        ws.cell(row=row_idx, column=1).value = code
        ws.cell(row=row_idx, column=2).value = desc
        ws.cell(row=row_idx, column=3).value = unit
        ws.cell(row=row_idx, column=4).value = price
    
    # Analytic sheet
    ws_ana = wb.create_sheet("Analitico-RS")
    ws_ana.cell(row=1, column=1).value = "CODIGO DA COMPOSICAO"
    ws_ana.cell(row=1, column=2).value = "ITEM DE COMPOSICAO"
    ws_ana.cell(row=1, column=3).value = "COEFICIENTE"
    
    ws_ana.cell(row=2, column=1).value = "50000"
    ws_ana.cell(row=2, column=2).value = "60000"
    ws_ana.cell(row=2, column=3).value = 10.0
    
    ws_ana.cell(row=3, column=1).value = "50000"
    ws_ana.cell(row=3, column=2).value = "60001"
    ws_ana.cell(row=3, column=3).value = 2.5
    
    ws_ana.cell(row=4, column=1).value = "50001"
    ws_ana.cell(row=4, column=2).value = "60000"
    ws_ana.cell(row=4, column=3).value = 5.0
    
    # Save
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    file_bytes = output.getvalue()
    
    # Zip
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zf:
        zf.writestr('SINAPI_RS_12_2025.xlsx', file_bytes)
    zip_buffer.seek(0)
    return zip_buffer.getvalue()

async def run_import():
    print("1. Importando SINAPI com composicoes...")
    file_bytes = await create_test_file()
    
    db = AsyncSessionLocal()
    try:
        result = await import_sinapi_excel(
            file_bytes,
            "SINAPI_RS_12_2025.zip",
            state="RS",
            month=12,
            year=2025,
            db=db
        )
        print(f"   [OK] Importado: {result} itens")
        
        # Get one service item ID
        item_q = await db.execute(text("""
            SELECT id FROM reference_items 
            WHERE code = '50000' AND source_id = (SELECT id FROM reference_sources WHERE name='SINAPI')
            LIMIT 1
        """))
        item_id = item_q.scalar()
        
        if not item_id:
            print("   [ERRO] Item nao encontrado!")
            return
        
        # Count compositions
        comp_q = await db.execute(text("SELECT COUNT(*) FROM composition_items"))
        comp_count = comp_q.scalar()
        print(f"   [OK] {comp_count} composition_items criadas")
        
        return str(item_id)
    finally:
        await db.close()

def test_api(item_id):
    print(f"\n2. Testando API com item {item_id[:8]}...")
    
    # Wait for server
    time.sleep(1)
    
    # Login
    try:
        r = requests.post("http://127.0.0.1:8000/api/v1/auth/token", 
                         data={"username": "admin@resolve.eng.br", "password": "admin123"},
                         timeout=5)
        if r.status_code != 200:
            print(f"   [ERRO] Login falhou: {r.status_code}")
            return False
        token = r.json()["access_token"]
        print("   [OK] Autenticado")
    except Exception as e:
        print(f"   [ERRO] Conexao falhou: {e}")
        print("   (servidor pode nao estar rodando - configure via FastAPI)")
        return None
    
    # Get composition
    try:
        headers = {"Authorization": f"Bearer {token}"}
        r = requests.get(f"http://127.0.0.1:8000/api/v1/catalog/composition/{item_id}?state=RS&charge_type=DESONERADO",
                        headers=headers, timeout=5)
        if r.status_code != 200:
            print(f"   [ERRO] Composicao nao encontrada: {r.status_code}")
            print(f"   Response: {r.text[:200]}")
            return False
        
        data = r.json()
        print(f"   [OK] Composicao recuperada para {data['code']}")
        print(f"       - {len(data['items'])} itens na composicao")
        for item in data['items']:
            print(f"         * {item['code']}: coef={item['coefficient']}, preco_unitario={item['unit_price']}")
        print(f"       - Custo calculado: R$ {data['calculated_cost']:.2f}")
        return True
    except Exception as e:
        print(f"   [ERRO] Teste da API falhou: {e}")
        return None

if __name__ == "__main__":
    print("=" * 60)
    print("TESTE COMPLETO: IMPORT + API")
    print("=" * 60)
    
    item_id = asyncio.run(run_import())
    if item_id:
        result = test_api(item_id)
        if result is True:
            print("\n" + "=" * 60)
            print("SUCESSO! Sistema funcionando corretamente.")
            print("=" * 60)
        elif result is False:
            print("\n[AVISO] API retornou erro, mas dados foram importados.")
        else:
            print("\n[INFO] Servidor nao esta rodando. Execute: uvicorn main:app --port 8000")
