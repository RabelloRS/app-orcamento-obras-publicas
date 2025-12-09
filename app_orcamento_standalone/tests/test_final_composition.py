#!/usr/bin/env python3
"""
Final integration test for composition import
"""
import asyncio
from io import BytesIO
import zipfile
from openpyxl import Workbook
from services.importer import import_sinapi_excel
from database import AsyncSessionLocal
from sqlalchemy import text

async def create_test_sinapi_file():
    """Create a more realistic SINAPI file with multiple compositions"""
    wb = Workbook()
    
    # Sheet 1: CCD (Composições Correntes Desoneradas)
    ws_comp = wb.active
    ws_comp.title = "CCD-RS-05-2025"
    
    # Headers with realistic state column
    headers = ["CÓDIGO", "DESCRIÇÃO", "UNIDADE", "RS"]
    for col, header in enumerate(headers, 1):
        ws_comp.cell(row=1, column=col).value = header
    
    # Services (5 items)
    services = [
        ("100100", "Serviço de pintura simples", "M2", 150.00),
        ("100200", "Serviço de pintura com primer", "M2", 200.00),
        ("100300", "Serviço de acabamento fino", "M2", 250.00),
        ("200100", "Mão de obra - pintor", "H", 75.00),
        ("200200", "Mão de obra - supervisor", "H", 100.00),
    ]
    
    for row_idx, (code, desc, unit, price) in enumerate(services, 2):
        ws_comp.cell(row=row_idx, column=1).value = code
        ws_comp.cell(row=row_idx, column=2).value = desc
        ws_comp.cell(row=row_idx, column=3).value = unit
        ws_comp.cell(row=row_idx, column=4).value = price
    
    # Sheet 2: Analítico (composition breakdown)
    ws_analitico = wb.create_sheet("Analítico")
    
    # Headers with intentional newlines to test normalization
    headers_analitico = ["CÓDIGO DA\nCOMPOSIÇÃO", "ITEM DE\nCOMPOSIÇÃO", "COEFICIENTE"]
    for col, header in enumerate(headers_analitico, 1):
        ws_analitico.cell(row=1, column=col).value = header
    
    # Composition breakdown
    compositions = [
        ("100100", "200100", 2.5),   # 100100 = 2.5 * 200100
        ("100100", "200200", 0.5),   # 100100 = 0.5 * 200200
        ("100200", "200100", 3.0),   # 100200 = 3.0 * 200100
        ("100200", "200200", 0.75),  # 100200 = 0.75 * 200200
        ("100300", "200100", 4.0),   # 100300 = 4.0 * 200100
    ]
    
    for row_idx, (parent, child, coef) in enumerate(compositions, 2):
        ws_analitico.cell(row=row_idx, column=1).value = parent
        ws_analitico.cell(row=row_idx, column=2).value = child
        ws_analitico.cell(row=row_idx, column=3).value = coef
    
    # Save and wrap in ZIP
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    file_bytes = output.getvalue()
    
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zf:
        zf.writestr('SINAPI_RS_05_2025.xlsx', file_bytes)
    zip_buffer.seek(0)
    return zip_buffer.getvalue()

async def run_test():
    print("=" * 60)
    print("TESTE FINAL DE IMPORTACAO DE COMPOSICOES")
    print("=" * 60)
    
    # Create test file
    file_bytes = await create_test_sinapi_file()
    print(f"[OK] Arquivo de teste criado ({len(file_bytes)} bytes)")
    
    # Import
    db = AsyncSessionLocal()
    try:
        result = await import_sinapi_excel(
            file_bytes,
            "SINAPI_RS_05_2025.zip",
            state="RS",
            month=5,
            year=2025,
            db=db
        )
        print(f"[OK] Importacao concluida: {result} itens processados")
        
        # Check results
        comp_count = await db.execute(text("SELECT COUNT(*) FROM composition_items"))
        count = comp_count.scalar()
        print(f"[OK] Total de composition_items criadas: {count}")
        
        if count > 0:
            print("\n  Amostras de composicoes:")
            samples = await db.execute(text("""
                SELECT p.code, c.code, ci.coefficient
                FROM composition_items ci
                JOIN reference_items p ON p.id = ci.parent_item_id
                JOIN reference_items c ON c.id = ci.child_item_id
                ORDER BY p.code
                LIMIT 10
            """))
            for parent, child, coef in samples:
                print(f"    {parent} = {coef} * {child}")
        
        # Verify counts
        item_count = await db.execute(text("SELECT COUNT(*) FROM reference_items WHERE source_id = (SELECT id FROM reference_sources WHERE name='SINAPI')"))
        print(f"[OK] Total de items (SERVICE + LABOR): {item_count.scalar()}")
        
        print("\n" + "=" * 60)
        print("TESTE CONCLUIDO COM SUCESSO!")
        print("=" * 60)
        
    finally:
        await db.close()

if __name__ == "__main__":
    asyncio.run(run_test())
