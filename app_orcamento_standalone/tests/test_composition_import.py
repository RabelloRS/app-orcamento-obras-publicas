#!/usr/bin/env python3
"""
Test composition import with simulated openpyxl workbook
"""
import asyncio
import sys
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO

async def test_composition_import():
    """
    Create a minimal SINAPI structure in memory and test composition import
    """
    # Create workbook
    wb = Workbook()
    
    # 1. Create Composição sheet (prices)
    ws_comp = wb.active
    ws_comp.title = "CCD-RS-05-2025"  # Composição Correntes Desoneradas
    
    # Headers - add state header
    headers = ["CÓDIGO", "DESCRIÇÃO", "UNIDADE", "RS", "COEFICIENTE TÉCNICO"]
    for col, header in enumerate(headers, 1):
        ws_comp.cell(row=1, column=col).value = header
    
    # Sample services
    ws_comp.cell(row=2, column=1).value = "001000"
    ws_comp.cell(row=2, column=2).value = "Serviço 001"
    ws_comp.cell(row=2, column=3).value = "UN"
    ws_comp.cell(row=2, column=4).value = 100.00  # RS price
    
    ws_comp.cell(row=3, column=1).value = "002000"
    ws_comp.cell(row=3, column=2).value = "Serviço 002"
    ws_comp.cell(row=3, column=3).value = "UN"
    ws_comp.cell(row=3, column=4).value = 200.00  # RS price
    
    # 2. Create Analítico sheet (composition breakdown)
    ws_analitico = wb.create_sheet("Analítico")
    
    # Headers with newlines to test normalization
    headers_analitico = ["CÓDIGO DA\nCOMPOSIÇÃO", "ITEM DE COMPOSIÇÃO", "COEFICIENTE"]
    for col, header in enumerate(headers_analitico, 1):
        ws_analitico.cell(row=1, column=col).value = header
    
    # Sample composition breakdown
    # Service 001000 decomposed into 002000 with coefficient 1.5
    ws_analitico.cell(row=2, column=1).value = "001000"
    ws_analitico.cell(row=2, column=2).value = "002000"
    ws_analitico.cell(row=2, column=3).value = 1.5
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    file_bytes = output.getvalue()
    
    # Wrap in ZIP (SINAPI expects ZIP format)
    import zipfile
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zf:
        zf.writestr('test_sinapi.xlsx', file_bytes)
    zip_buffer.seek(0)
    file_bytes = zip_buffer.getvalue()
    
    # Now test the import
    print("[TEST] Saved test workbook")
    
    # Import it
    from services.importer import import_sinapi_excel
    from database import AsyncSessionLocal
    from models import ReferenceItem, CompositionItem
    
    db = AsyncSessionLocal()
    try:
        result = await import_sinapi_excel(
            file_bytes, 
            "test_sinapi_202505.xlsx",
            state="RS",
            month=5,
            year=2025,
            db=db
        )
        print(f"[TEST] Import result: {result}")
        
        # Check database
        from sqlalchemy import text
        
        # Count services
        service_count = await db.execute(text("SELECT COUNT(*) FROM reference_items WHERE type='SERVICE'"))
        print(f"[TEST] Total SERVICE items: {service_count.scalar()}")
        
        # Count compositions
        comp_count = await db.execute(text("SELECT COUNT(*) FROM composition_items"))
        print(f"[TEST] Total composition_items: {comp_count.scalar()}")
        
        # List compositions
        comps = await db.execute(text("SELECT * FROM composition_items LIMIT 10"))
        for row in comps:
            print(f"[TEST] Composition: {row}")
    finally:
        await db.close()

if __name__ == "__main__":
    asyncio.run(test_composition_import())
